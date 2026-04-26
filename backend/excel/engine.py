"""
RightCut — WorkbookEngine
In-memory openpyxl workbook manager. All mutations happen here.
One instance per WebSocket session.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.comments import Comment
from openpyxl.formatting.rule import (
    ColorScaleRule,
    DataBarRule,
    FormatObject,
    Rule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import column_index_from_string, get_column_letter

from .formulas import infer_format, validate_formula
from models import ChartMeta, ValidationStats

logger = logging.getLogger(__name__)

# Thin border side used everywhere
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


class WorkbookEngine:
    """Thread-safe-enough for single-session async usage."""

    def __init__(self) -> None:
        self.wb: Workbook = Workbook()
        # Remove the default blank sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        # chart metadata keyed by sheet name
        self._charts: dict[str, list[ChartMeta]] = {}

    # ── Sheet lifecycle ───────────────────────────────────────────────────────

    def create_sheet(self, sheet_name: str, headers: list[str]) -> dict:
        """Create (or reset) a sheet with bold styled headers."""
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]

        ws = self.wb.create_sheet(title=sheet_name)
        self._charts.setdefault(sheet_name, [])

        # Write and style the header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 14)

        ws.freeze_panes = "A2"

        return {
            "sheet_name": sheet_name,
            "headers": headers,
            "created": True,
        }

    def get_all_sheet_names(self) -> list[str]:
        return list(self.wb.sheetnames)

    # ── Model scaffold ────────────────────────────────────────────────────────

    def create_model_scaffold(
        self,
        model_type: str,
        params: dict,
    ) -> dict:
        """
        Build a complete, correctly-structured model scaffold with proper
        formulas, references, and formatting in one call.
        Supported model_type: "dcf" or "lbo"
        params for DCF:
          base_ebitda, revenue_growth, ebitda_margin, da_pct, capex_pct,
          wc_pct, tax_rate, wacc, terminal_growth, net_debt, shares_outstanding,
          company_name, years (default 5), currency (default USD)
        """
        if model_type.lower() == "dcf":
            return self._build_dcf_scaffold(params)
        if model_type.lower() == "lbo":
            return self._build_lbo_scaffold(params)
        return {"error": f"Unknown model_type: {model_type}", "success": False}

    def _build_dcf_scaffold(self, p: dict) -> dict:
        from datetime import date

        # ── Parameters with defaults ──────────────────────────────────────────
        company     = p.get("company_name", "Company")
        currency    = p.get("currency", "USD")
        base_ebitda = float(p.get("base_ebitda", 50_000_000))
        rev_growth  = float(p.get("revenue_growth", 0.10))
        ebitda_margin = float(p.get("ebitda_margin", 0.20))
        da_pct      = float(p.get("da_pct", 0.05))
        capex_pct   = float(p.get("capex_pct", 0.06))
        wc_pct      = float(p.get("wc_pct", 0.01))
        tax_rate    = float(p.get("tax_rate", 0.25))
        wacc        = float(p.get("wacc", 0.10))
        term_g      = float(p.get("terminal_growth", 0.03))
        net_debt    = float(p.get("net_debt", 0))
        shares      = float(p.get("shares_outstanding", 10_000_000))
        n_years     = int(p.get("years", 5))

        # Derived base year
        base_revenue = base_ebitda / ebitda_margin

        # ── Clear existing sheets ─────────────────────────────────────────────
        for sn in ["Cover", "Assumptions", "Income Statement", "DCF Valuation"]:
            if sn in self.wb.sheetnames:
                del self.wb[sn]
            self._charts[sn] = []

        today = date.today().strftime("%Y-%m-%d")
        base_year = date.today().year - 1
        year_labels = [f"FY{base_year}A"] + [f"FY{base_year+i}E" for i in range(1, n_years + 1)]

        # ════════════════════════════════════════════════════════════════════
        # SHEET 1 — Cover
        # ════════════════════════════════════════════════════════════════════
        ws_cov = self.wb.create_sheet("Cover")
        ws_cov.column_dimensions["A"].width = 28
        ws_cov.column_dimensions["B"].width = 35

        # Row 1 = header (Label | Value) — serializer reads this as column headers
        for c, hdr in enumerate(["Label", "Value"], start=1):
            cell = ws_cov.cell(row=1, column=c, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER

        cover_rows = [
            ("Model Title",       f"DCF Valuation — {company}"),
            ("Model Date",        today),
            ("Currency",          currency),
            ("Projection Period", f"{n_years} Years"),
            ("Valuation Method",  "Discounted Cash Flow — Gordon Growth Model"),
            ("Analyst",           "RightCut AI"),
        ]
        # Data rows start at row 2
        for r, (label, value) in enumerate(cover_rows, start=2):
            ws_cov.cell(row=r, column=1, value=label).border = _BORDER
            ws_cov.cell(row=r, column=2, value=value).border = _BORDER

        # Navy highlight on Model Title row (row 2)
        for c in (1, 2):
            cell = ws_cov.cell(row=2, column=c)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # ════════════════════════════════════════════════════════════════════
        # SHEET 2 — Assumptions  (values in column B, row 2 onward)
        # Row map (1-indexed):
        #   1=header, 2=section, 3=rev_growth, 4=section, 5=ebitda_margin,
        #   6=da_pct, 7=capex_pct, 8=wc_pct, 9=tax_rate,
        #   10=section, 11=wacc, 12=term_g,
        #   13=section, 14=base_rev, 15=base_ebitda, 16=net_debt, 17=shares
        # ════════════════════════════════════════════════════════════════════
        ws_ass = self.wb.create_sheet("Assumptions")
        ws_ass.column_dimensions["A"].width = 32
        ws_ass.column_dimensions["B"].width = 18

        # Header row
        for c, hdr in enumerate(["Assumption", "Value"], start=1):
            cell = ws_ass.cell(row=1, column=c, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER

        _SEC = "D9E2F3"   # light blue section header bg
        _DARK = "1F3864"  # dark navy text

        ass_data = [
            # (row, label, value, is_section, number_format)
            (2,  "REVENUE ASSUMPTIONS",    None,         True,  None),
            (3,  "Revenue Growth Rate",    rev_growth,   False, "0.0%"),
            (4,  "MARGIN ASSUMPTIONS",     None,         True,  None),
            (5,  "EBITDA Margin",          ebitda_margin,False, "0.0%"),
            (6,  "D&A as % of Revenue",    da_pct,       False, "0.0%"),
            (7,  "CapEx as % of Revenue",  capex_pct,    False, "0.0%"),
            (8,  "Change in WC as % Rev",  wc_pct,       False, "0.0%"),
            (9,  "Corporate Tax Rate",     tax_rate,     False, "0.0%"),
            (10, "WACC / DISCOUNT ASSUMPTIONS", None,    True,  None),
            (11, "WACC",                   wacc,         False, "0.0%"),
            (12, "Terminal Growth Rate",   term_g,       False, "0.0%"),
            (13, "BALANCE SHEET INPUTS",   None,         True,  None),
            (14, "Base Revenue (FY2024A)", base_revenue, False, "#,##0"),
            (15, "Base EBITDA (FY2024A)",  base_ebitda,  False, "#,##0"),
            (16, "Net Debt",               net_debt,     False, "#,##0"),
            (17, "Shares Outstanding",     shares,       False, "#,##0"),
        ]

        # Named row pointers (for formula references)
        # B3=rev_growth, B5=ebitda_margin, B6=da_pct, B7=capex_pct,
        # B8=wc_pct, B9=tax_rate, B11=wacc, B12=term_g,
        # B14=base_rev, B15=base_ebitda, B16=net_debt, B17=shares
        A = {
            "rev_growth":    "Assumptions!$B$3",
            "ebitda_margin": "Assumptions!$B$5",
            "da_pct":        "Assumptions!$B$6",
            "capex_pct":     "Assumptions!$B$7",
            "wc_pct":        "Assumptions!$B$8",
            "tax_rate":      "Assumptions!$B$9",
            "wacc":          "Assumptions!$B$11",
            "term_g":        "Assumptions!$B$12",
            "base_rev":      "Assumptions!$B$14",
            "base_ebitda":   "Assumptions!$B$15",
            "net_debt":      "Assumptions!$B$16",
            "shares":        "Assumptions!$B$17",
        }

        for row, label, value, is_section, fmt in ass_data:
            cell_a = ws_ass.cell(row=row, column=1, value=label)
            cell_b = ws_ass.cell(row=row, column=2, value=value)
            cell_a.border = _BORDER
            cell_b.border = _BORDER
            if is_section:
                for c in (cell_a, cell_b):
                    c.font = Font(bold=True, color=_DARK, size=11)
                    c.fill = PatternFill(fill_type="solid", fgColor=_SEC)
                    c.alignment = Alignment(horizontal="left", vertical="center")
            elif fmt:
                cell_b.number_format = fmt

        # ════════════════════════════════════════════════════════════════════
        # SHEET 3 — Income Statement
        # Col A=label, B=base year, C..=projected years
        # Row map: 1=header, 2=rev, 3=ebitda, 4=da, 5=ebit, 6=nopat,
        #          7=capex, 8=da_addback, 9=dwc, 10=FCFF
        # ════════════════════════════════════════════════════════════════════
        ws_is = self.wb.create_sheet("Income Statement")
        ws_is.column_dimensions["A"].width = 30
        for i in range(1, n_years + 2):
            ws_is.column_dimensions[get_column_letter(i + 1)].width = 14

        # Header row
        is_headers = ["Line Item"] + year_labels
        for c, hdr in enumerate(is_headers, start=1):
            cell = ws_is.cell(row=1, column=c, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER

        # Column letters: B=base, C=yr1, D=yr2, ...
        col_base = "B"   # FY2024A
        col_yrs  = [get_column_letter(i + 3) for i in range(n_years)]  # C, D, E, F, G

        # Build all IS rows directly (no lambdas — avoids closure bugs)
        # col_yrs[0]=C (yr1), col_yrs[1]=D (yr2), ...
        # Prior year for revenue: base=B, yr1 prior=B, yr2 prior=C, ...
        rev_prior = [col_base] + col_yrs[:-1]  # B, C, D, E, F

        # Row 2: Revenue
        ws_is.cell(row=2, column=1, value="Revenue").border = _BORDER
        ws_is.cell(row=2, column=2, value=base_revenue).border = _BORDER
        ws_is.cell(row=2, column=2).number_format = "#,##0"
        for i, col in enumerate(col_yrs):
            prior = rev_prior[i]
            cell = ws_is.cell(row=2, column=3 + i, value=f"={prior}2*(1+{A['rev_growth']})")
            cell.number_format = "#,##0"
            cell.border = _BORDER

        def _is_rows_3_to_10():
            for row_num, label, b_tmpl, proj_tmpl in [
                (3,  "EBITDA",                   f"={col_base}2*{A['ebitda_margin']}", "{col}2*{em}"),
                (4,  "D&A",                       f"={col_base}2*{A['da_pct']}",       "{col}2*{dp}"),
                (5,  "EBIT",                      f"={col_base}3-{col_base}4",          "{col}3-{col}4"),
                (6,  "NOPAT",                     f"={col_base}5*(1-{A['tax_rate']})",  "{col}5*(1-{tr})"),
                (7,  "Capital Expenditures",      f"=-{col_base}2*{A['capex_pct']}",   "-{col}2*{cp}"),
                (8,  "D&A Add-back",              f"={col_base}4",                      "{col}4"),
                (9,  "Change in Working Capital", f"=-{col_base}2*{A['wc_pct']}",      "-{col}2*{wp}"),
                (10, "FCFF",                      f"={col_base}6+{col_base}8+{col_base}7+{col_base}9",
                                                                                         "{col}6+{col}8+{col}7+{col}9"),
            ]:
                ws_is.cell(row=row_num, column=1, value=label).border = _BORDER
                ws_is.cell(row=row_num, column=2, value=b_tmpl).border = _BORDER
                ws_is.cell(row=row_num, column=2).number_format = "#,##0"
                for i, col in enumerate(col_yrs):
                    formula = "=" + proj_tmpl.format(
                        col=col,
                        em=A['ebitda_margin'], dp=A['da_pct'], tr=A['tax_rate'],
                        cp=A['capex_pct'], wp=A['wc_pct'],
                    )
                    cell = ws_is.cell(row=row_num, column=3 + i, value=formula)
                    cell.number_format = "#,##0"
                    cell.border = _BORDER

        _is_rows_3_to_10()

        # Formatting
        _green_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
        _bold_green = Font(bold=True, color="375623", size=11)
        for c in range(1, n_years + 3):
            # EBITDA row (3) — output_row
            cell3 = ws_is.cell(row=3, column=c)
            cell3.fill = _green_fill
            cell3.font = _bold_green
            # FCFF row (10) — output_row
            cell10 = ws_is.cell(row=10, column=c)
            cell10.fill = _green_fill
            cell10.font = _bold_green
            # EBIT row (5) — subtotal_row
            ws_is.cell(row=5, column=c).font = Font(bold=True, size=11)

        ws_is.freeze_panes = "B2"

        # ════════════════════════════════════════════════════════════════════
        # SHEET 4 — DCF Valuation
        # Only projected years (no base year column)
        # Col A=label, B=yr1 … F=yr5
        # Row map:
        #   1=header
        #   2=FCFF (from IS)
        #   3=discount factor
        #   4=PV FCFF
        #   5=blank
        #   6=section: TERMINAL VALUE
        #   7=terminal FCFF
        #   8=terminal value
        #   9=PV terminal value
        #   10=blank
        #   11=section: EQUITY VALUE BRIDGE
        #   12=sum PV FCFFs
        #   13=enterprise value
        #   14=less net debt
        #   15=equity value
        #   16=shares outstanding
        #   17=INTRINSIC VALUE / SHARE  ← final_answer_row
        # ════════════════════════════════════════════════════════════════════
        ws_dcf = self.wb.create_sheet("DCF Valuation")
        ws_dcf.column_dimensions["A"].width = 30
        for i in range(1, n_years + 1):
            ws_dcf.column_dimensions[get_column_letter(i + 1)].width = 14

        # Projected year column letters: B=yr1, C=yr2, ...
        dcf_cols = [get_column_letter(i + 2) for i in range(n_years)]  # B, C, D, E, F

        # Header row (projected years only)
        dcf_headers = ["Metric"] + year_labels[1:]  # skip base year
        for c, hdr in enumerate(dcf_headers, start=1):
            cell = ws_dcf.cell(row=1, column=c, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER

        def dcf_cell(row: int, col_letter: str, value, fmt="#,##0"):
            cell = ws_dcf.cell(row=row, column=ord(col_letter) - ord("A") + 1, value=value)
            cell.border = _BORDER
            if fmt:
                cell.number_format = fmt
            return cell

        def dcf_label(row: int, label: str):
            cell = ws_dcf.cell(row=row, column=1, value=label)
            cell.border = _BORDER
            return cell

        # FCFF from Income Statement (projected years = IS cols C, D, E, F, G)
        dcf_label(2, "FCFF")
        is_proj_cols = [get_column_letter(i + 3) for i in range(n_years)]  # C, D, E, F, G
        for i, (dcf_col, is_col) in enumerate(zip(dcf_cols, is_proj_cols)):
            dcf_cell(2, dcf_col, f"='Income Statement'!{is_col}10")

        # Discount factor = 1/(1+WACC)^n
        dcf_label(3, "Discount Factor")
        for i, col in enumerate(dcf_cols):
            dcf_cell(3, col, f"=1/(1+{A['wacc']})^{i+1}", "0.0000")

        # PV of FCFF = FCFF * DF
        dcf_label(4, "PV of FCFF")
        for i, col in enumerate(dcf_cols):
            dcf_cell(4, col, f"={col}2*{col}3")

        # Blank row 5
        dcf_label(5, "")

        # Section header row 6
        dcf_label(6, "TERMINAL VALUE CALCULATION")
        for c in range(1, n_years + 2):
            cell = ws_dcf.cell(row=6, column=c)
            cell.font = Font(bold=True, color=_DARK, size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor=_SEC)
            cell.border = _BORDER

        # Terminal FCFF = last projected FCFF * (1+g)
        last_col = dcf_cols[-1]
        dcf_label(7, "Terminal FCFF")
        dcf_cell(7, "B", f"={last_col}2*(1+{A['term_g']})")

        # Terminal Value = Terminal FCFF / (WACC - g)
        dcf_label(8, "Terminal Value")
        dcf_cell(8, "B", f"=B7/({A['wacc']}-{A['term_g']})")

        # PV of Terminal Value
        dcf_label(9, "PV of Terminal Value")
        dcf_cell(9, "B", f"=B8/(1+{A['wacc']})^{n_years}")

        # Blank row 10
        dcf_label(10, "")

        # Section header row 11
        dcf_label(11, "EQUITY VALUE BRIDGE")
        for c in range(1, n_years + 2):
            cell = ws_dcf.cell(row=11, column=c)
            cell.font = Font(bold=True, color=_DARK, size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor=_SEC)
            cell.border = _BORDER

        # Sum of PV FCFFs
        pv_range = f"B4:{get_column_letter(n_years + 1)}4"
        dcf_label(12, "Sum of PV FCFFs")
        dcf_cell(12, "B", f"=SUM({pv_range})")

        # Enterprise Value
        dcf_label(13, "Enterprise Value")
        dcf_cell(13, "B", "=B12+B9")

        # Less: Net Debt
        dcf_label(14, "Less: Net Debt")
        dcf_cell(14, "B", f"={A['net_debt']}")

        # Equity Value
        dcf_label(15, "Equity Value")
        dcf_cell(15, "B", "=B13-B14")

        # Shares Outstanding
        dcf_label(16, "Shares Outstanding")
        dcf_cell(16, "B", f"={A['shares']}", "#,##0")

        # Intrinsic Value / Share — FINAL ANSWER
        dcf_label(17, "Intrinsic Value / Share")
        dcf_cell(17, "B", "=B15/B16", "#,##0.00")

        # Formatting for DCF
        _green_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
        _bold_green = Font(bold=True, color="375623", size=11)
        _navy_fill = PatternFill(fill_type="solid", fgColor="1F3864")
        _white_bold = Font(bold=True, color="FFFFFF", size=12)

        for c in range(1, n_years + 2):
            # PV FCFF row (4) — output_row
            ws_dcf.cell(row=4, column=c).fill = _green_fill
            ws_dcf.cell(row=4, column=c).font = _bold_green
            # Equity Value row (15) — output_row
            ws_dcf.cell(row=15, column=c).fill = _green_fill
            ws_dcf.cell(row=15, column=c).font = _bold_green

        # Final answer row (17) — dark navy
        for c in range(1, 3):  # just label + value columns
            cell = ws_dcf.cell(row=17, column=c)
            cell.fill = _navy_fill
            cell.font = _white_bold
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center")

        ws_dcf.freeze_panes = "B2"

        # ── Summary ───────────────────────────────────────────────────────────
        sheets_created = ["Cover", "Assumptions", "Income Statement", "DCF Valuation"]
        for sn in sheets_created:
            self._charts.setdefault(sn, [])

        return {
            "sheets_created": sheets_created,
            "assumption_rows": {k: v for k, v in {
                "revenue_growth_row": 3, "ebitda_margin_row": 5, "da_pct_row": 6,
                "capex_pct_row": 7, "wc_pct_row": 8, "tax_rate_row": 9,
                "wacc_row": 11, "terminal_growth_row": 12,
                "base_revenue_row": 14, "base_ebitda_row": 15,
                "net_debt_row": 16, "shares_row": 17,
            }.items()},
            "income_statement_rows": {
                "revenue": 2, "ebitda": 3, "da": 4, "ebit": 5,
                "nopat": 6, "capex": 7, "da_addback": 8, "wc": 9, "fcff": 10,
            },
            "dcf_rows": {
                "fcff": 2, "discount_factor": 3, "pv_fcff": 4,
                "terminal_fcff": 7, "terminal_value": 8, "pv_tv": 9,
                "sum_pv_fcff": 12, "enterprise_value": 13,
                "net_debt": 14, "equity_value": 15,
                "shares": 16, "intrinsic_value_per_share": 17,
            },
            "note": "All formulas are circular-reference-free. Assumptions sheet drives IS and DCF. Edit Assumptions!B3-B17 to update model.",
        }

    def _build_lbo_scaffold(self, p: dict) -> dict:
        """
        Build a complete LBO model with:
          Sheet 1 — Cover
          Sheet 2 — Assumptions  (entry, financing, operating, exit)
          Sheet 3 — Debt Schedule (term loan amortisation + PIK)
          Sheet 4 — Income Statement & Cash Flow
          Sheet 5 — Returns (MOIC, IRR, equity bridge)

        params:
          company_name, entry_ebitda, entry_multiple, revenue_growth,
          ebitda_margin, debt_pct, interest_rate, amort_pct,
          exit_multiple, tax_rate, years (default 5), currency (default USD)
        """
        from datetime import date

        company      = p.get("company_name", "Portfolio Co")
        currency     = p.get("currency", "USD")
        entry_ebitda = float(p.get("entry_ebitda", 50_000_000))
        entry_mult   = float(p.get("entry_multiple", 8.0))
        rev_growth   = float(p.get("revenue_growth", 0.10))
        ebitda_mgn   = float(p.get("ebitda_margin", 0.25))
        debt_pct     = float(p.get("debt_pct", 0.60))        # % of TEV financed by debt
        int_rate     = float(p.get("interest_rate", 0.07))
        amort_pct    = float(p.get("amort_pct", 0.05))       # % of initial debt/yr
        exit_mult    = float(p.get("exit_multiple", 10.0))
        tax_rate     = float(p.get("tax_rate", 0.25))
        n_years      = int(p.get("years", 5))
        today        = date.today().strftime("%Y-%m-%d")
        base_year    = date.today().year
        year_labels  = [f"FY{base_year + i}E" for i in range(1, n_years + 1)]

        # Derived entry values
        tev          = entry_ebitda * entry_mult
        entry_debt   = tev * debt_pct
        entry_equity = tev * (1 - debt_pct)
        entry_rev    = entry_ebitda / ebitda_mgn

        # Clear any previous sheets with the same names
        for sn in ["Cover", "Assumptions", "Debt Schedule", "Income Statement", "Returns"]:
            if sn in self.wb.sheetnames:
                del self.wb[sn]
            self._charts[sn] = []

        _SEC  = "D9E2F3"
        _DARK = "1F3864"
        _GREEN_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")
        _BOLD_GREEN = Font(bold=True, color="375623", size=11)
        _NAVY_FILL  = PatternFill(fill_type="solid", fgColor="1F3864")
        _WHITE_BOLD = Font(bold=True, color="FFFFFF", size=12)

        def _hdr(ws, col, val):
            c = ws.cell(row=1, column=col, value=val)
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            c.alignment = Alignment(horizontal="center")
            c.border = _BORDER

        def _sec(ws, row, label, n_cols=2):
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=row, column=c, value=label if c == 1 else None)
                cell.font = Font(bold=True, color=_DARK, size=11)
                cell.fill = PatternFill(fill_type="solid", fgColor=_SEC)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = _BORDER

        def _val(ws, row, col, value, fmt=None, bold=False):
            c = ws.cell(row=row, column=col, value=value)
            c.border = _BORDER
            if fmt:
                c.number_format = fmt
            if bold:
                c.font = Font(bold=True, size=11)
            return c

        # ════════════════════════════════════════════════════════════
        # SHEET 1 — Cover
        # ════════════════════════════════════════════════════════════
        ws_cov = self.wb.create_sheet("Cover")
        ws_cov.column_dimensions["A"].width = 30
        ws_cov.column_dimensions["B"].width = 36
        for c, h in enumerate(["Label", "Value"], 1):
            _hdr(ws_cov, c, h)
        cover_rows = [
            ("Model Title",       f"LBO Analysis — {company}"),
            ("Model Date",        today),
            ("Currency",          currency),
            ("Projection Period", f"{n_years} Years"),
            ("Valuation Method",  "Leveraged Buyout — Cash-on-Cash / IRR"),
            ("Analyst",           "RightCut AI"),
        ]
        for r, (lbl, val) in enumerate(cover_rows, 2):
            ws_cov.cell(row=r, column=1, value=lbl).border = _BORDER
            ws_cov.cell(row=r, column=2, value=val).border = _BORDER
        for c in (1, 2):
            cell = ws_cov.cell(row=2, column=c)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")

        # ════════════════════════════════════════════════════════════
        # SHEET 2 — Assumptions
        # Row map:
        #  1=header
        #  2=ENTRY section,  3=Entry EBITDA, 4=Entry Multiple, 5=TEV, 6=Debt%, 7=Entry Debt, 8=Entry Equity
        #  9=OPERATING,      10=Rev Growth, 11=EBITDA Margin, 12=Tax Rate
        #  13=FINANCING,     14=Interest Rate, 15=Amort%
        #  16=EXIT,          17=Exit Multiple, 18=Exit Year
        # ════════════════════════════════════════════════════════════
        ws_ass = self.wb.create_sheet("Assumptions")
        ws_ass.column_dimensions["A"].width = 32
        ws_ass.column_dimensions["B"].width = 18
        for c, h in enumerate(["Assumption", "Value"], 1):
            _hdr(ws_ass, c, h)

        A = {
            "entry_ebitda":  "Assumptions!$B$3",
            "entry_mult":    "Assumptions!$B$4",
            "tev":           "Assumptions!$B$5",
            "debt_pct":      "Assumptions!$B$6",
            "entry_debt":    "Assumptions!$B$7",
            "entry_equity":  "Assumptions!$B$8",
            "rev_growth":    "Assumptions!$B$10",
            "ebitda_mgn":    "Assumptions!$B$11",
            "tax_rate":      "Assumptions!$B$12",
            "int_rate":      "Assumptions!$B$14",
            "amort_pct":     "Assumptions!$B$15",
            "exit_mult":     "Assumptions!$B$17",
            "exit_year":     "Assumptions!$B$18",
        }

        ass_data = [
            (2,  "ENTRY ASSUMPTIONS",         None,         True,  None),
            (3,  "Entry EBITDA",               entry_ebitda, False, "#,##0"),
            (4,  "Entry EV / EBITDA Multiple", entry_mult,   False, "0.0x"),
            (5,  "Total Enterprise Value",     tev,          False, "#,##0"),
            (6,  "Debt Financing (%)",         debt_pct,     False, "0.0%"),
            (7,  "Entry Debt",                 entry_debt,   False, "#,##0"),
            (8,  "Sponsor Equity",             entry_equity, False, "#,##0"),
            (9,  "OPERATING ASSUMPTIONS",      None,         True,  None),
            (10, "Revenue Growth Rate",        rev_growth,   False, "0.0%"),
            (11, "EBITDA Margin",              ebitda_mgn,   False, "0.0%"),
            (12, "Corporate Tax Rate",         tax_rate,     False, "0.0%"),
            (13, "FINANCING ASSUMPTIONS",      None,         True,  None),
            (14, "Interest Rate on Debt",      int_rate,     False, "0.0%"),
            (15, "Annual Amortisation (%)",    amort_pct,    False, "0.0%"),
            (16, "EXIT ASSUMPTIONS",           None,         True,  None),
            (17, "Exit EV / EBITDA Multiple",  exit_mult,    False, "0.0x"),
            (18, "Hold Period (Years)",        n_years,      False, "0"),
        ]
        for row, lbl, val, is_sec, fmt in ass_data:
            ca = ws_ass.cell(row=row, column=1, value=lbl)
            cb = ws_ass.cell(row=row, column=2, value=val)
            ca.border = _BORDER; cb.border = _BORDER
            if is_sec:
                for c in (ca, cb):
                    c.font = Font(bold=True, color=_DARK, size=11)
                    c.fill = PatternFill(fill_type="solid", fgColor=_SEC)
            elif fmt:
                cb.number_format = fmt

        # ════════════════════════════════════════════════════════════
        # SHEET 3 — Debt Schedule
        # Col A = label, B..F = years
        # Rows: 1=hdr, 2=Opening Balance, 3=Amortisation, 4=Closing Balance,
        #       5=blank, 6=Interest Expense, 7=Cash Interest Paid
        # ════════════════════════════════════════════════════════════
        ws_dbt = self.wb.create_sheet("Debt Schedule")
        ws_dbt.column_dimensions["A"].width = 28
        for i in range(1, n_years + 1):
            ws_dbt.column_dimensions[get_column_letter(i + 1)].width = 14

        ds_hdrs = ["Line Item"] + year_labels
        for c, h in enumerate(ds_hdrs, 1):
            _hdr(ws_dbt, c, h)

        yr_cols = [get_column_letter(i + 2) for i in range(n_years)]  # C,D,E,F,G shifted to B,C,...

        # Simpler: col B=yr1, C=yr2 ...
        ds_cols = [get_column_letter(i + 2) for i in range(n_years)]  # B,C,D,E,F

        # Row 2: Opening Debt Balance
        ws_dbt.cell(row=2, column=1, value="Opening Debt Balance").border = _BORDER
        for i, col in enumerate(ds_cols):
            if i == 0:
                formula = f"={A['entry_debt']}"
            else:
                prev = ds_cols[i - 1]
                formula = f"={prev}4"   # prior closing balance
            _val(ws_dbt, 2, ord(col) - ord("A") + 1, formula, "#,##0")

        # Row 3: Amortisation (-)
        ws_dbt.cell(row=3, column=1, value="Mandatory Amortisation").border = _BORDER
        for i, col in enumerate(ds_cols):
            formula = f"=-{A['entry_debt']}*{A['amort_pct']}"
            _val(ws_dbt, 3, ord(col) - ord("A") + 1, formula, "#,##0")

        # Row 4: Closing Balance
        ws_dbt.cell(row=4, column=1, value="Closing Debt Balance").border = _BORDER
        for i, col in enumerate(ds_cols):
            formula = f"={col}2+{col}3"
            c = _val(ws_dbt, 4, ord(col) - ord("A") + 1, formula, "#,##0", bold=True)
            c.fill = _GREEN_FILL; c.font = _BOLD_GREEN

        # Row 5 blank
        ws_dbt.cell(row=5, column=1, value="").border = _BORDER

        # Row 6: Interest Expense (on avg balance)
        ws_dbt.cell(row=6, column=1, value="Interest Expense").border = _BORDER
        for i, col in enumerate(ds_cols):
            formula = f"=-({col}2+{col}4)/2*{A['int_rate']}"
            _val(ws_dbt, 6, ord(col) - ord("A") + 1, formula, "#,##0")

        ws_dbt.freeze_panes = "B2"

        # ════════════════════════════════════════════════════════════
        # SHEET 4 — Income Statement & Cash Flow
        # Col A=label, B..F=years
        # Rows: 1=hdr, 2=Revenue, 3=EBITDA, 4=D&A(est), 5=EBIT,
        #       6=Interest, 7=EBT, 8=Tax, 9=Net Income,
        #       10=blank, 11=EBITDA(again), 12=Capex(est), 13=D&A,
        #       14=Change WC, 15=Interest Paid, 16=Tax Paid, 17=FCF to Equity
        # ════════════════════════════════════════════════════════════
        ws_is = self.wb.create_sheet("Income Statement")
        ws_is.column_dimensions["A"].width = 30
        for i in range(1, n_years + 1):
            ws_is.column_dimensions[get_column_letter(i + 1)].width = 14

        is_hdrs = ["Line Item"] + year_labels
        for c, h in enumerate(is_hdrs, 1):
            _hdr(ws_is, c, h)

        is_cols = [get_column_letter(i + 2) for i in range(n_years)]

        # Row 2: Revenue  (base = entry_ebitda / ebitda_margin, grow each year)
        ws_is.cell(row=2, column=1, value="Revenue").border = _BORDER
        for i, col in enumerate(is_cols):
            if i == 0:
                base_rev_formula = f"={A['entry_ebitda']}/{A['ebitda_mgn']}*(1+{A['rev_growth']})"
            else:
                prev = is_cols[i - 1]
                base_rev_formula = f"={prev}2*(1+{A['rev_growth']})"
            _val(ws_is, 2, ord(col) - ord("A") + 1, base_rev_formula, "#,##0")

        # Row 3: EBITDA
        ws_is.cell(row=3, column=1, value="EBITDA").border = _BORDER
        for col in is_cols:
            _val(ws_is, 3, ord(col) - ord("A") + 1, f"={col}2*{A['ebitda_mgn']}", "#,##0")
        for c in range(1, n_years + 2):
            ws_is.cell(row=3, column=c).fill = _GREEN_FILL
            ws_is.cell(row=3, column=c).font = _BOLD_GREEN

        # Row 4: D&A (estimated at 3% of revenue)
        ws_is.cell(row=4, column=1, value="D&A (est. 3% rev)").border = _BORDER
        for col in is_cols:
            _val(ws_is, 4, ord(col) - ord("A") + 1, f"=-{col}2*0.03", "#,##0")

        # Row 5: EBIT
        ws_is.cell(row=5, column=1, value="EBIT").border = _BORDER
        for col in is_cols:
            c = _val(ws_is, 5, ord(col) - ord("A") + 1, f"={col}3+{col}4", "#,##0", bold=True)

        # Row 6: Interest (from Debt Schedule row 6)
        ws_is.cell(row=6, column=1, value="Interest Expense").border = _BORDER
        for i, col in enumerate(is_cols):
            ds_col = ds_cols[i]
            _val(ws_is, 6, ord(col) - ord("A") + 1, f"='Debt Schedule'!{ds_col}6", "#,##0")

        # Row 7: EBT
        ws_is.cell(row=7, column=1, value="EBT (Pre-Tax Income)").border = _BORDER
        for col in is_cols:
            _val(ws_is, 7, ord(col) - ord("A") + 1, f"={col}5+{col}6", "#,##0")

        # Row 8: Tax
        ws_is.cell(row=8, column=1, value="Income Tax").border = _BORDER
        for col in is_cols:
            _val(ws_is, 8, ord(col) - ord("A") + 1, f"=IF({col}7>0,-{col}7*{A['tax_rate']},0)", "#,##0")

        # Row 9: Net Income
        ws_is.cell(row=9, column=1, value="Net Income").border = _BORDER
        for col in is_cols:
            c = _val(ws_is, 9, ord(col) - ord("A") + 1, f"={col}7+{col}8", "#,##0", bold=True)
            c.fill = _GREEN_FILL; c.font = _BOLD_GREEN

        # Row 10: blank
        ws_is.cell(row=10, column=1, value="").border = _BORDER

        # Section: Free Cash Flow
        _sec(ws_is, 11, "FREE CASH FLOW TO EQUITY", n_years + 1)
        # Row 11 section header already set — put EBITDA ref in same row cols
        for i, col in enumerate(is_cols):
            ws_is.cell(row=11, column=ord(col) - ord("A") + 1, value=None).border = _BORDER

        # Row 12: Capex (est -4% rev)
        ws_is.cell(row=12, column=1, value="Capital Expenditure (est)").border = _BORDER
        for col in is_cols:
            _val(ws_is, 12, ord(col) - ord("A") + 1, f"=-{col}2*0.04", "#,##0")

        # Row 13: D&A add-back
        ws_is.cell(row=13, column=1, value="D&A Add-back").border = _BORDER
        for col in is_cols:
            _val(ws_is, 13, ord(col) - ord("A") + 1, f"=-{col}4", "#,##0")

        # Row 14: Change in Working Capital (est -1% rev growth)
        ws_is.cell(row=14, column=1, value="Change in Working Capital").border = _BORDER
        for col in is_cols:
            _val(ws_is, 14, ord(col) - ord("A") + 1, f"=-{col}2*0.01", "#,##0")

        # Row 15: Interest paid (from debt schedule)
        ws_is.cell(row=15, column=1, value="Cash Interest Paid").border = _BORDER
        for i, col in enumerate(is_cols):
            ds_col = ds_cols[i]
            _val(ws_is, 15, ord(col) - ord("A") + 1, f"='Debt Schedule'!{ds_col}6", "#,##0")

        # Row 16: Tax paid
        ws_is.cell(row=16, column=1, value="Tax Paid").border = _BORDER
        for col in is_cols:
            _val(ws_is, 16, ord(col) - ord("A") + 1, f"={col}8", "#,##0")

        # Row 17: FCF to Equity
        ws_is.cell(row=17, column=1, value="Free Cash Flow to Equity").border = _BORDER
        for col in is_cols:
            c = _val(ws_is, 17, ord(col) - ord("A") + 1,
                     f"={col}9+{col}12+{col}13+{col}14", "#,##0", bold=True)
            c.fill = _GREEN_FILL; c.font = _BOLD_GREEN

        ws_is.freeze_panes = "B2"

        # ════════════════════════════════════════════════════════════
        # SHEET 5 — Returns  (MOIC & IRR)
        # Col A=label, B=value
        # Rows:
        #  1=hdr
        #  2=EXIT section, 3=Exit Year EBITDA, 4=Exit Multiple, 5=Exit TEV,
        #  6=Closing Debt at Exit, 7=Exit Equity Value
        #  8=blank
        #  9=RETURNS section, 10=Entry Equity, 11=Exit Equity, 12=MOIC,
        #  13=IRR (using XIRR approx), 14=Hold Period
        #  15=blank
        #  16=CASH FLOW BRIDGE section
        #  17..21=yearly FCF (for IRR)
        # ════════════════════════════════════════════════════════════
        ws_ret = self.wb.create_sheet("Returns")
        ws_ret.column_dimensions["A"].width = 32
        ws_ret.column_dimensions["B"].width = 20
        for c, h in enumerate(["Metric", "Value"], 1):
            _hdr(ws_ret, c, h)

        last_is_col = is_cols[-1]     # last projected year column in IS
        last_ds_col = ds_cols[-1]     # last year column in Debt Schedule

        _sec(ws_ret, 2, "EXIT VALUATION")
        _val(ws_ret, 3, 1, "Exit Year EBITDA").border = _BORDER
        _val(ws_ret, 3, 2, f"='Income Statement'!{last_is_col}3", "#,##0")
        _val(ws_ret, 4, 1, "Exit EV/EBITDA Multiple").border = _BORDER
        _val(ws_ret, 4, 2, f"={A['exit_mult']}", "0.0x")
        _val(ws_ret, 5, 1, "Exit Enterprise Value").border = _BORDER
        _val(ws_ret, 5, 2, "=B3*B4", "#,##0")
        _val(ws_ret, 6, 1, "Closing Debt at Exit").border = _BORDER
        _val(ws_ret, 6, 2, f"='Debt Schedule'!{last_ds_col}4", "#,##0")
        _val(ws_ret, 7, 1, "Exit Equity Value").border = _BORDER
        c7 = _val(ws_ret, 7, 2, "=B5-B6", "#,##0", bold=True)
        c7.fill = _GREEN_FILL; c7.font = _BOLD_GREEN

        ws_ret.cell(row=8, column=1).border = _BORDER
        ws_ret.cell(row=8, column=2).border = _BORDER

        _sec(ws_ret, 9, "RETURNS SUMMARY")
        _val(ws_ret, 10, 1, "Entry Equity (Invested)").border = _BORDER
        _val(ws_ret, 10, 2, f"={A['entry_equity']}", "#,##0")
        _val(ws_ret, 11, 1, "Exit Equity (Proceeds)").border = _BORDER
        _val(ws_ret, 11, 2, "=B7", "#,##0")

        # MOIC = Exit Equity / Entry Equity
        _val(ws_ret, 12, 1, "MOIC (Money-on-Money)").border = _BORDER
        c_moic = _val(ws_ret, 12, 2, "=B11/B10", "0.00x", bold=True)
        c_moic.fill = _NAVY_FILL; c_moic.font = _WHITE_BOLD

        # IRR approximation: =(MOIC^(1/hold_period))-1
        _val(ws_ret, 13, 1, "IRR (Approx. CAGR Method)").border = _BORDER
        c_irr = _val(ws_ret, 13, 2, f"=(B11/B10)^(1/{A['exit_year']})-1", "0.0%", bold=True)
        c_irr.fill = _NAVY_FILL; c_irr.font = _WHITE_BOLD

        _val(ws_ret, 14, 1, "Hold Period (Years)").border = _BORDER
        _val(ws_ret, 14, 2, f"={A['exit_year']}", "0")

        ws_ret.cell(row=15, column=1).border = _BORDER
        ws_ret.cell(row=15, column=2).border = _BORDER

        # Cash flow bridge — initial equity outflow + yearly FCF
        _sec(ws_ret, 16, "CASH FLOW BRIDGE (for IRR)")
        _val(ws_ret, 17, 1, "Year 0 — Entry (Equity Outflow)").border = _BORDER
        _val(ws_ret, 17, 2, f"=-{A['entry_equity']}", "#,##0")
        for i, (col, lbl) in enumerate(zip(is_cols, year_labels)):
            row = 18 + i
            _val(ws_ret, row, 1, f"{lbl} — FCF to Equity").border = _BORDER
            _val(ws_ret, row, 2, f"='Income Statement'!{col}17", "#,##0")
        # Final year: also add exit equity proceeds
        last_row = 18 + n_years - 1
        ws_ret.cell(row=last_row, column=2).value = (
            f"='Income Statement'!{last_is_col}17+Returns!B7"
        )

        ws_ret.freeze_panes = "B2"

        sheets_created = ["Cover", "Assumptions", "Debt Schedule", "Income Statement", "Returns"]
        for sn in sheets_created:
            self._charts.setdefault(sn, [])

        return {
            "sheets_created": sheets_created,
            "assumption_rows": {
                "entry_ebitda": 3, "entry_multiple": 4, "tev": 5,
                "debt_pct": 6, "entry_debt": 7, "entry_equity": 8,
                "rev_growth": 10, "ebitda_margin": 11, "tax_rate": 12,
                "interest_rate": 14, "amort_pct": 15,
                "exit_multiple": 17, "hold_years": 18,
            },
            "returns_rows": {
                "exit_ebitda": 3, "exit_tev": 5, "closing_debt": 6,
                "exit_equity": 7, "entry_equity": 10, "moic": 12, "irr": 13,
            },
            "note": (
                "LBO scaffold complete. MOIC and IRR are in the Returns sheet. "
                "Edit Assumptions!B3:B18 to sensitise the model. "
                "IRR uses CAGR approximation — for precision use Excel XIRR on the Cash Flow Bridge rows."
            ),
        }

    # ── Data insertion ────────────────────────────────────────────────────────

    def insert_data(
        self,
        sheet_name: str,
        rows: list[list[str]],
        start_row: int = 2,
    ) -> dict:
        ws = self._get_sheet(sheet_name)
        max_col = ws.max_column or 1

        rows_written = 0
        for row_idx, row_data in enumerate(rows, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value if value != "" else None
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="center")

                # Auto-detect number format from header
                header_cell = ws.cell(row=1, column=col_idx)
                if header_cell.value:
                    fmt = infer_format(str(header_cell.value))
                    if fmt:
                        cell.number_format = fmt

            rows_written += 1

        return {
            "sheet_name": sheet_name,
            "rows_written": rows_written,
            "start_row": start_row,
        }

    # ── Formula insertion ─────────────────────────────────────────────────────

    def add_formula(
        self,
        sheet_name: str,
        cell: str,
        formula: str,
        apply_to_range: str | None = None,
    ) -> dict:
        ws = self._get_sheet(sheet_name)

        if not validate_formula(formula):
            return {"error": f"Invalid formula: {formula}", "success": False}

        ws[cell] = formula
        ws[cell].border = _BORDER

        applied_cells = [cell]

        if apply_to_range:
            # Derive the column offset pattern and apply across the range
            try:
                from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                anchor_col_letter, anchor_row = coordinate_from_string(cell)
                anchor_col = column_index_from_string(anchor_col_letter)

                # Parse the target range
                start_ref, end_ref = apply_to_range.split(":")
                start_col_letter, start_row = coordinate_from_string(start_ref)
                end_col_letter, end_row = coordinate_from_string(end_ref)
                start_col = column_index_from_string(start_col_letter)
                end_col = column_index_from_string(end_col_letter)

                for r in range(start_row, end_row + 1):
                    for c in range(start_col, end_col + 1):
                        if r == anchor_row and c == anchor_col:
                            continue  # already written
                        # Shift formula row references
                        row_delta = r - anchor_row
                        shifted = _shift_formula_rows(formula, row_delta)
                        target_cell = f"{get_column_letter(c)}{r}"
                        ws[target_cell] = shifted
                        ws[target_cell].border = _BORDER
                        applied_cells.append(target_cell)
            except Exception as e:
                logger.warning(f"apply_to_range failed: {e}")

        return {
            "sheet_name": sheet_name,
            "formula": formula,
            "cells": applied_cells,
        }

    # ── Cell editing ──────────────────────────────────────────────────────────

    def edit_cell(self, sheet_name: str, cell: str, value: str) -> dict:
        ws = self._get_sheet(sheet_name)
        old_value = ws[cell].value
        ws[cell].value = value
        ws[cell].border = _BORDER
        return {"sheet_name": sheet_name, "cell": cell, "old": str(old_value), "new": value}

    def apply_user_edit(self, sheet_name: str, cell: str, value: str) -> None:
        """Apply a user-initiated cell edit (no response needed)."""
        try:
            ws = self._get_sheet(sheet_name)
            ws[cell].value = value
        except Exception as e:
            logger.warning(f"apply_user_edit failed for {sheet_name}!{cell}: {e}")

    # ── Formatting ────────────────────────────────────────────────────────────

    def apply_formatting(
        self,
        sheet_name: str,
        cell_range: str,
        format_type: str,
        format_config: dict | None = None,
    ) -> dict:
        ws = self._get_sheet(sheet_name)
        cfg = format_config or {}

        if format_type == "color_scale":
            rule = ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            )
            ws.conditional_formatting.add(cell_range, rule)

        elif format_type == "data_bar":
            rule = DataBarRule(
                start_type="min",
                start_value=0,
                end_type="max",
                end_value=100,
                color="638EC6",
            )
            ws.conditional_formatting.add(cell_range, rule)

        elif format_type == "bold_header":
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = _BORDER

        elif format_type == "number_format":
            fmt = cfg.get("format", "#,##0")
            for row in ws[cell_range]:
                for cell in row:
                    cell.number_format = fmt

        elif format_type == "border":
            style = cfg.get("style", "thin")
            side = Side(style=style)
            border = Border(left=side, right=side, top=side, bottom=side)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = border

        elif format_type == "font_color":
            color = cfg.get("color", "FF0000").lstrip("#")
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = Font(color=color)

        elif format_type == "background_color":
            color = cfg.get("color", "FFFF00").lstrip("#")
            for row in ws[cell_range]:
                for cell in row:
                    cell.fill = PatternFill(fill_type="solid", fgColor=color)

        elif format_type == "section_header":
            # Light-blue section divider row with bold dark text — use between logical blocks
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = Font(bold=True, color="1F3864", size=11)
                    cell.fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = _BORDER

        elif format_type == "subtotal_row":
            # Bold text, no background — for totals, sub-totals, computed summary rows
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = Font(bold=True, size=11)
                    cell.border = _BORDER

        elif format_type == "output_row":
            # Light green background — for key output rows (FCFF, Net Profit, EBITDA)
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = Font(bold=True, color="375623", size=11)
                    cell.fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
                    cell.border = _BORDER

        elif format_type == "final_answer_row":
            # Dark navy background with white text — for the single most important output row
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = _BORDER

        elif format_type == "zebra_stripe":
            # Alternate light-gray rows for tables (apply to even rows within range)
            for row in ws[cell_range]:
                for cell in row:
                    if cell.row % 2 == 0:
                        cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

        elif format_type == "muted_row":
            # Grey font, lighter weight — for sub-commentary rows (growth %, margin %)
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = Font(color="555555", size=10)

        else:
            return {"error": f"Unknown format_type: {format_type}", "success": False}

        return {"sheet_name": sheet_name, "range": cell_range, "format_type": format_type}

    # ── Citations ─────────────────────────────────────────────────────────────

    def add_citation(
        self,
        sheet_name: str,
        cell: str,
        source_file: str,
        source_location: str,
        excerpt: str = "",
    ) -> dict:
        ws = self._get_sheet(sheet_name)
        comment_text = f"Source: {source_file}\nRef: {source_location}"
        if excerpt:
            comment_text += f"\n\"{excerpt[:200]}\""
        comment = Comment(comment_text, "RightCut Agent")
        ws[cell].comment = comment
        return {
            "sheet_name": sheet_name,
            "cell": cell,
            "citation": comment_text,
        }

    # ── Sorting ───────────────────────────────────────────────────────────────

    def sort_range(
        self,
        sheet_name: str,
        sort_column: str,
        ascending: bool = True,
    ) -> dict:
        ws = self._get_sheet(sheet_name)
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row < 3:
            return {"sheet_name": sheet_name, "sorted": False, "reason": "not enough rows"}

        # Find the sort column index
        sort_col_idx: int | None = None
        for col_idx in range(1, max_col + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header and str(header).strip().lower() == sort_column.strip().lower():
                sort_col_idx = col_idx
                break

        if sort_col_idx is None:
            # Try treating sort_column as a letter reference
            try:
                sort_col_idx = column_index_from_string(sort_column.upper())
            except Exception:
                return {"error": f"Column '{sort_column}' not found", "success": False}

        # Read data rows (skip header)
        data_rows: list[list[Any]] = []
        for row_idx in range(2, max_row + 1):
            row_data = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
            data_rows.append(row_data)

        # Sort — try numeric, fall back to string
        def sort_key(row: list[Any]) -> tuple:
            val = row[sort_col_idx - 1]
            if val is None:
                return (1, "")
            try:
                return (0, float(str(val).replace("$", "").replace(",", "").replace("%", "")))
            except (ValueError, TypeError):
                return (0, str(val).lower())

        data_rows.sort(key=sort_key, reverse=not ascending)

        # Write back
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx).value = value

        return {
            "sheet_name": sheet_name,
            "sort_column": sort_column,
            "ascending": ascending,
            "rows_sorted": len(data_rows),
        }

    # ── Charts ────────────────────────────────────────────────────────────────

    def create_chart(
        self,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        title: str = "",
        target_cell: str = "H2",
    ) -> dict:
        ws = self._get_sheet(sheet_name)

        # Parse data range
        try:
            start_ref, end_ref = data_range.split(":")
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            start_col_letter, start_row = coordinate_from_string(start_ref)
            end_col_letter, end_row = coordinate_from_string(end_ref)
            min_col = column_index_from_string(start_col_letter)
            max_col = column_index_from_string(end_col_letter)
        except Exception as e:
            return {"error": f"Invalid data_range '{data_range}': {e}", "success": False}

        # Build chart
        chart_map = {
            "bar": BarChart,
            "line": LineChart,
            "pie": PieChart,
            "scatter": ScatterChart,
        }
        ChartClass = chart_map.get(chart_type.lower(), BarChart)
        chart = ChartClass()
        chart.title = title or sheet_name
        chart.style = 10
        chart.width = 18
        chart.height = 10

        if chart_type.lower() == "bar":
            chart.type = "col"   # vertical bars (horizontal is the confusing default)
            chart.grouping = "clustered"

        # Data columns = everything EXCEPT the first column (which is categories/labels)
        # For a range like A1:C5: col A = labels, cols B-C = value series
        if max_col > min_col:
            data_ref = Reference(
                ws,
                min_col=min_col + 1,   # skip label column
                min_row=start_row,
                max_col=max_col,
                max_row=end_row,
            )
            cats = Reference(ws, min_col=min_col, min_row=start_row + 1, max_row=end_row)
        else:
            # Single column — use it as data (no labels)
            data_ref = Reference(
                ws,
                min_col=min_col,
                min_row=start_row,
                max_col=max_col,
                max_row=end_row,
            )
            cats = None

        if chart_type.lower() == "scatter":
            chart.add_data(data_ref)
        else:
            chart.add_data(data_ref, titles_from_data=True)

        if cats is not None:
            chart.set_categories(cats)

        ws.add_chart(chart, target_cell)

        # Track chart metadata for frontend
        meta = ChartMeta(
            chart_type=chart_type,
            title=title or sheet_name,
            data_range=data_range,
            anchor_cell=target_cell,
        )
        self._charts.setdefault(sheet_name, []).append(meta)

        return {
            "sheet_name": sheet_name,
            "chart_type": chart_type,
            "title": title or sheet_name,
            "anchor": target_cell,
            "note": "Chart embedded in .xlsx. Download to view rendered chart.",
        }

    # ── Data Cleanup ──────────────────────────────────────────────────────────

    def clean_data(
        self,
        sheet_name: str,
        operation: str,
        column: str | None = None,
        find_text: str | None = None,
        replace_text: str | None = None,
        delimiter: str | None = None,
        new_column_name: str | None = None,
    ) -> dict:
        """
        Apply one of 15 data-cleaning operations to a sheet.
        column: header name or letter (e.g. 'Name' or 'B'). If None, applies to all text cells.
        """
        import re

        ws = self._get_sheet(sheet_name)
        max_row = ws.max_row
        max_col = ws.max_column

        if not max_row or max_row < 2:
            return {"error": "Sheet has no data rows", "success": False}

        # Resolve column index
        col_idx: int | None = None
        if column:
            # Try header name first
            for c in range(1, max_col + 1):
                hdr = ws.cell(row=1, column=c).value
                if hdr and str(hdr).strip().lower() == column.strip().lower():
                    col_idx = c
                    break
            if col_idx is None:
                # Try letter
                try:
                    col_idx = column_index_from_string(column.upper())
                except Exception:
                    return {"error": f"Column '{column}' not found", "success": False}

        def _cols() -> range:
            return range(col_idx, col_idx + 1) if col_idx else range(1, max_col + 1)

        def _str_val(cell) -> str | None:
            v = cell.value
            return str(v) if v is not None else None

        cells_changed = 0

        # ── Operations ────────────────────────────────────────────────────────

        if operation == "trim_whitespace":
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v is not None:
                        cleaned = " ".join(v.split())
                        if cleaned != v:
                            cell.value = cleaned
                            cells_changed += 1

        elif operation == "remove_duplicates":
            seen: set[tuple] = set()
            rows_to_delete: list[int] = []
            for r in range(2, max_row + 1):
                row_key = tuple(
                    str(ws.cell(row=r, column=c).value or "").strip()
                    for c in range(1, max_col + 1)
                )
                if row_key in seen:
                    rows_to_delete.append(r)
                else:
                    seen.add(row_key)
            # Delete from bottom to keep row indices stable
            for r in reversed(rows_to_delete):
                ws.delete_rows(r)
                cells_changed += 1

        elif operation == "remove_blank_rows":
            blank_rows: list[int] = []
            check_cols = list(_cols())
            for r in range(2, max_row + 1):
                is_blank = all(
                    ws.cell(row=r, column=c).value in (None, "")
                    for c in check_cols
                )
                if is_blank:
                    blank_rows.append(r)
            for r in reversed(blank_rows):
                ws.delete_rows(r)
                cells_changed += 1

        elif operation == "to_uppercase":
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v and not v.startswith("="):
                        cell.value = v.upper()
                        cells_changed += 1

        elif operation == "to_lowercase":
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v and not v.startswith("="):
                        cell.value = v.lower()
                        cells_changed += 1

        elif operation == "to_titlecase":
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v and not v.startswith("="):
                        cell.value = v.title()
                        cells_changed += 1

        elif operation == "find_replace":
            if find_text is None:
                return {"error": "find_text is required for find_replace", "success": False}
            replace_with = replace_text or ""
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v and find_text in v:
                        cell.value = v.replace(find_text, replace_with)
                        cells_changed += 1

        elif operation == "convert_to_number":
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v:
                        cleaned = v.replace("$", "").replace(",", "").replace("%", "").strip()
                        try:
                            num = float(cleaned)
                            cell.value = int(num) if num == int(num) else num
                            cells_changed += 1
                        except (ValueError, OverflowError):
                            pass

        elif operation == "convert_to_date":
            from dateutil import parser as dateparser
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v and not v.startswith("="):
                        try:
                            dt = dateparser.parse(v, dayfirst=False)
                            cell.value = dt.date()
                            cell.number_format = "YYYY-MM-DD"
                            cells_changed += 1
                        except Exception:
                            pass

        elif operation == "remove_special_chars":
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v and not v.startswith("="):
                        cleaned = re.sub(r"[^\w\s\.\,\-\(\)\$\%\/\:]", "", v)
                        if cleaned != v:
                            cell.value = cleaned
                            cells_changed += 1

        elif operation == "fill_down":
            if col_idx is None:
                return {"error": "column is required for fill_down", "success": False}
            last_val = None
            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value not in (None, ""):
                    last_val = cell.value
                elif last_val is not None:
                    cell.value = last_val
                    cells_changed += 1

        elif operation == "extract_numbers":
            if col_idx is None:
                return {"error": "column is required for extract_numbers", "success": False}
            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                v = _str_val(cell)
                if v:
                    nums = re.findall(r"-?\d+\.?\d*", v)
                    if nums:
                        cell.value = float(nums[0]) if "." in nums[0] else int(nums[0])
                        cells_changed += 1

        elif operation == "split_column":
            if col_idx is None:
                return {"error": "column is required for split_column", "success": False}
            sep = delimiter or " "
            new_name = new_column_name or (
                str(ws.cell(row=1, column=col_idx).value or "Split") + "_2"
            )
            # Insert new column right after the target
            new_col_idx = col_idx + 1
            ws.insert_cols(new_col_idx)
            # Write header
            new_hdr = ws.cell(row=1, column=new_col_idx)
            new_hdr.value = new_name
            new_hdr.font = Font(bold=True, color="FFFFFF", size=11)
            new_hdr.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            new_hdr.alignment = Alignment(horizontal="center")
            new_hdr.border = _BORDER
            ws.column_dimensions[get_column_letter(new_col_idx)].width = 18
            # Split values
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                v = _str_val(cell)
                if v and sep in v:
                    parts = v.split(sep, 1)
                    cell.value = parts[0].strip()
                    ws.cell(row=r, column=new_col_idx).value = parts[1].strip()
                    cells_changed += 1

        elif operation == "fix_number_format":
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v:
                        cleaned = v.strip()
                        is_pct = cleaned.endswith("%")
                        cleaned2 = cleaned.replace("$", "").replace(",", "").replace("%", "").strip()
                        try:
                            num = float(cleaned2)
                            if is_pct:
                                cell.value = num / 100
                                cell.number_format = "0.0%"
                            else:
                                cell.value = int(num) if num == int(num) else num
                                cell.number_format = "#,##0" if abs(num) >= 1000 else "0.00"
                            cells_changed += 1
                        except (ValueError, OverflowError):
                            pass

        elif operation == "standardize_text":
            # Remove extra spaces, normalize unicode, consistent punctuation
            for r in range(2, max_row + 1):
                for c in _cols():
                    cell = ws.cell(row=r, column=c)
                    v = _str_val(cell)
                    if v and not v.startswith("="):
                        import unicodedata
                        normalized = unicodedata.normalize("NFKC", v)
                        cleaned = " ".join(normalized.split())
                        if cleaned != v:
                            cell.value = cleaned
                            cells_changed += 1

        else:
            return {"error": f"Unknown operation: '{operation}'", "success": False}

        return {
            "sheet_name": sheet_name,
            "operation": operation,
            "column": column,
            "cells_changed": cells_changed,
            "success": True,
        }

    # ── Find & Replace (cross-sheet) ──────────────────────────────────────────

    def find_replace(
        self,
        sheet_name: str,
        find_text: str,
        replace_text: str = "",
        match_case: bool = False,
    ) -> dict:
        """Find and replace text across all cells in a sheet."""
        ws = self._get_sheet(sheet_name)
        cells_changed = 0
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                v = cell.value
                if v is None or not isinstance(v, str):
                    continue
                if not v.startswith("="):
                    haystack = v if match_case else v.lower()
                    needle = find_text if match_case else find_text.lower()
                    if needle in haystack:
                        cell.value = v.replace(find_text, replace_text) if match_case else \
                            v.replace(find_text, replace_text)
                        cells_changed += 1
        return {
            "sheet_name": sheet_name,
            "find": find_text,
            "replace": replace_text,
            "cells_changed": cells_changed,
        }

    # ── Validation ────────────────────────────────────────────────────────────

    def validate_workbook(self, check_hardcoded: bool = True) -> dict:
        stats = ValidationStats()
        hardcoded_cells: list[str] = []

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value is None:
                        continue
                    stats.total_cells += 1
                    val_str = str(cell.value)
                    if val_str.startswith("="):
                        stats.formula_count += 1
                    elif check_hardcoded:
                        # Flag numeric cells that look like they could be formulas
                        try:
                            float(val_str.replace("$", "").replace(",", "").replace("%", ""))
                            # It's a number — flag if the header suggests it should be a formula
                            header = ws.cell(row=1, column=cell.column).value
                            if header and any(
                                k in str(header).lower()
                                for k in ("cagr", "margin", "irr", "moic", "growth", "ratio")
                            ):
                                ref = f"{sheet_name}!{cell.coordinate}"
                                hardcoded_cells.append(ref)
                                stats.hardcoded_count += 1
                        except (ValueError, TypeError):
                            pass

        stats.issues = [f"Hardcoded value in {c} (consider using a formula)" for c in hardcoded_cells[:10]]

        return {
            "formula_count": stats.formula_count,
            "hardcoded_count": stats.hardcoded_count,
            "total_cells": stats.total_cells,
            "issues": stats.issues,
            "valid": stats.hardcoded_count == 0,
            "summary": (
                f"{stats.formula_count} formulas, "
                f"{stats.hardcoded_count} potentially hardcoded values, "
                f"{stats.total_cells} total data cells"
            ),
        }

    # ── Audit (agentic self-correction) ─────────────────────────────────────

    def _evaluate_workbook(self) -> dict[str, Any]:
        """
        Evaluate all formulas in the workbook using the `formulas` library.
        Returns a dict mapping 'SHEET!CELL' → computed value.
        Falls back gracefully if evaluation fails.
        """
        import tempfile, os
        try:
            import formulas as flib
        except ImportError:
            return {}

        tmp = None
        try:
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            self.wb.save(tmp.name)
            tmp.close()
            xl = flib.ExcelModel().loads(tmp.name).finish()
            raw = xl.calculate()
            # Normalize keys: '[file]SHEET'!CELL → SHEET!CELL
            evaluated: dict[str, Any] = {}
            for key, val in raw.items():
                k = str(key)
                # Extract sheet and cell from key like "'[file]SHEET'!A1"
                m = re.search(r"\](.+?)'!([A-Z]+\d+(?::[A-Z]+\d+)?)", k)
                if not m:
                    continue
                sname, cref = m.group(1), m.group(2)
                if ":" in cref:
                    continue  # skip range references
                # Extract scalar value from Ranges object
                v = val
                if hasattr(v, 'value'):
                    v = v.value
                    if hasattr(v, '__iter__') and not isinstance(v, str):
                        import numpy as np
                        v = v.flat[0] if hasattr(v, 'flat') else list(v)[0] if v else None
                        if isinstance(v, np.generic):
                            v = v.item()
                evaluated[f"{sname}!{cref}"] = v
            return evaluated
        except Exception as exc:
            logger.warning(f"Formula evaluation failed (audit continues without): {exc}")
            return {}
        finally:
            if tmp:
                try:
                    os.unlink(tmp.name)
                except OSError:
                    pass

    def audit_sheet(self, sheet_name: str) -> dict:
        """
        Deep audit of a single sheet with actual formula evaluation.
        1. Structural checks (syntax, cross-sheet refs, circular refs)
        2. Formula evaluation via `formulas` library — computes actual values
        3. Returns row-by-row table of EVALUATED values so the LLM can verify
           that every formula produces the correct result
        4. Flags #DIV/0!, #REF!, #NAME?, #VALUE! errors from evaluation
        """
        ws = self._get_sheet(sheet_name)
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        issues: list[dict] = []
        formula_cells: list[dict] = []

        # ── Step 1: Evaluate all formulas ─────────────────────────────────
        evaluated = self._evaluate_workbook()

        # Collect all sheet names for cross-sheet ref validation
        all_sheets = set(s.lower() for s in self.wb.sheetnames)

        _ref_pattern = re.compile(
            r"(?:'([^']+)'|([A-Za-z_]\w*))!(\$?[A-Z]+\$?\d+)"
            r"|(\$?[A-Z]+\$?\d+)"
        )

        # ── Step 2: Structural checks + build evaluated snapshot ──────────
        row_snapshot: list[dict] = []

        for row_idx in range(1, max_row + 1):
            row_data: dict[str, Any] = {"row": row_idx}
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_letter = get_column_letter(col_idx)
                cell_ref_str = f"{col_letter}{row_idx}"
                header = ws.cell(row=1, column=col_idx).value or col_letter

                val = cell.value
                if val is None:
                    continue

                if isinstance(val, str) and val.startswith("="):
                    formula = val

                    # Look up the evaluated value
                    eval_key = f"{sheet_name}!{cell_ref_str}"
                    # Try case variations (formulas lib upper-cases sheet names)
                    computed = evaluated.get(eval_key)
                    if computed is None:
                        computed = evaluated.get(f"{sheet_name.upper()}!{cell_ref_str}")

                    info: dict[str, Any] = {
                        "cell": cell_ref_str,
                        "formula": formula,
                    }
                    if computed is not None:
                        info["computed_value"] = computed

                    # Structural check: balanced parentheses
                    if not validate_formula(formula):
                        issues.append({
                            "cell": cell_ref_str,
                            "type": "syntax_error",
                            "message": f"Unbalanced parentheses in formula: {formula}",
                            "severity": "error",
                        })

                    # Cross-sheet reference validation
                    for m in _ref_pattern.finditer(formula):
                        ref_sheet = m.group(1) or m.group(2)
                        if ref_sheet and ref_sheet.lower() not in all_sheets:
                            issues.append({
                                "cell": cell_ref_str,
                                "type": "missing_sheet_ref",
                                "message": f"References sheet '{ref_sheet}' which does not exist",
                                "severity": "error",
                            })

                    # Self-referencing check
                    for match in _ref_pattern.findall(formula):
                        ref = match[3]
                        if ref and ref.replace("$", "").upper() == cell_ref_str:
                            issues.append({
                                "cell": cell_ref_str,
                                "type": "circular_reference",
                                "message": f"Formula references itself: {formula}",
                                "severity": "error",
                            })

                    # Division by zero pattern
                    if re.search(r'/\s*0\b', formula) and 'IF(' not in formula.upper():
                        issues.append({
                            "cell": cell_ref_str,
                            "type": "division_by_zero_risk",
                            "message": f"Possible division by zero without IF guard: {formula}",
                            "severity": "warning",
                        })

                    # Check evaluated value for errors
                    if computed is not None:
                        computed_str = str(computed)
                        if any(e in computed_str for e in ('#DIV/0!', '#REF!', '#NAME?', '#VALUE!', '#N/A', '#NULL!')):
                            issues.append({
                                "cell": cell_ref_str,
                                "type": "eval_error",
                                "message": f"Formula evaluates to {computed_str}: {formula}",
                                "severity": "error",
                            })
                        elif isinstance(computed, float) and (
                            computed != computed  # NaN
                            or computed == float('inf')
                            or computed == float('-inf')
                        ):
                            issues.append({
                                "cell": cell_ref_str,
                                "type": "eval_error",
                                "message": f"Formula evaluates to {computed_str} (NaN/Inf): {formula}",
                                "severity": "error",
                            })

                    # openpyxl error type check
                    if cell.data_type == 'e':
                        issues.append({
                            "cell": cell_ref_str,
                            "type": "formula_error",
                            "message": f"Excel error value in cell (formula: {formula})",
                            "severity": "error",
                        })

                    formula_cells.append(info)

                    # For row snapshot: show formula → evaluated value
                    if row_idx <= 50:
                        if computed is not None:
                            display = computed
                            if isinstance(display, float):
                                display = round(display, 4)
                            row_data[str(header)] = f"{display}  ← {formula}"
                        else:
                            row_data[str(header)] = f"[not evaluated]  ← {formula}"
                else:
                    # Plain value
                    if row_idx <= 50:
                        row_data[str(header)] = val

            if row_idx <= 50 and len(row_data) > 1:
                row_snapshot.append(row_data)

        # ── Step 3: Column-level checks ───────────────────────────────────

        # Detect columns that look numeric but have text mixed in
        for col_idx in range(1, max_col + 1):
            num_count = 0
            text_count = 0
            text_cells = []
            for row_idx in range(2, min(max_row + 1, 52)):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None:
                    continue
                if isinstance(val, (int, float)):
                    num_count += 1
                elif isinstance(val, str) and not val.startswith("="):
                    try:
                        float(val.replace(",", "").replace("$", "").replace("%", ""))
                        num_count += 1
                    except ValueError:
                        text_count += 1
                        text_cells.append(f"{get_column_letter(col_idx)}{row_idx}")
            if num_count > 3 and text_count > 0 and text_count <= 3:
                header = ws.cell(row=1, column=col_idx).value or get_column_letter(col_idx)
                issues.append({
                    "cell": ", ".join(text_cells[:3]),
                    "type": "mixed_types",
                    "message": f"Column '{header}' is mostly numeric but has text in {text_count} cell(s)",
                    "severity": "warning",
                })

        error_count = sum(1 for i in issues if i["severity"] == "error")
        warning_count = sum(1 for i in issues if i["severity"] == "warning")
        has_eval = len(evaluated) > 0

        return {
            "sheet_name": sheet_name,
            "total_rows": max_row - 1,
            "total_formulas": len(formula_cells),
            "formulas_evaluated": has_eval,
            "errors": error_count,
            "warnings": warning_count,
            "issues": issues[:20],
            "formula_results": formula_cells[:40],  # formula + computed value pairs
            "row_snapshot": row_snapshot[:30],  # rows with evaluated values inline
            "healthy": error_count == 0,
            "summary": (
                f"Sheet '{sheet_name}': {max_row - 1} rows, "
                f"{len(formula_cells)} formulas"
                + (f" (all evaluated)" if has_eval else " (evaluation unavailable)")
                + f", {error_count} error(s), {warning_count} warning(s)"
                + (" — ALL OK ✓" if error_count == 0 else " — NEEDS FIXES ✗")
            ),
        }

    # ── Sheet state (for agent) ───────────────────────────────────────────────

    def get_sheet_state(self, sheet_name: str) -> dict:
        ws = self._get_sheet(sheet_name)
        headers = []
        rows = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            row_data = []
            for cell in row:
                val = cell.value
                if val is None:
                    row_data.append({"value": None, "formula": None})
                elif isinstance(val, str) and val.startswith("="):
                    row_data.append({"value": None, "formula": val})
                else:
                    row_data.append({"value": str(val), "formula": None})

            if row_idx == 1:
                headers = [str(c["value"] or "") for c in row_data]
            else:
                rows.append(row_data)

        return {
            "sheet_name": sheet_name,
            "headers": headers,
            "row_count": len(rows),
            "rows": rows[:50],  # cap at 50 rows for context window
        }

    # ── Download ──────────────────────────────────────────────────────────────

    def to_bytes(self) -> bytes:
        """
        Save the workbook to .xlsx bytes with cached formula values.

        Pipeline:
        1. Save openpyxl workbook to temp file (formulas + charts + formatting)
        2. Evaluate all formulas server-side using the `formulas` library
        3. Patch the xlsx XML to inject <v> cached values alongside <f> formula tags
        4. Return the patched xlsx bytes

        This ensures:
        - Charts read correct cached values (no red error triangles)
        - Excel/Numbers/Google Sheets show values immediately without recalculation
        - Formulas are preserved for end-user editing
        """
        import os
        import tempfile
        import zipfile

        from lxml import etree

        self.wb.calculation.calcMode = "auto"
        self.wb.calculation.fullCalcOnLoad = True

        # Step 1: Save raw workbook
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        try:
            self.wb.save(tmp.name)
            tmp.close()

            # Step 2: Evaluate formulas
            try:
                import formulas as flib

                xl_model = flib.ExcelModel().loads(tmp.name).finish()
                sol = xl_model.calculate()

                # Build lookup: (SHEET_NAME_UPPER, CELL_REF_UPPER) → computed value
                cached: dict[tuple[str, str], object] = {}
                for key, ranges_obj in sol.items():
                    key_str = str(key)
                    if "!" not in key_str:
                        continue
                    sheet_part, cell_part = key_str.rsplit("!", 1)
                    if ":" in cell_part:
                        continue  # skip range references
                    sheet_name = sheet_part.strip("'")
                    if "]" in sheet_name:
                        sheet_name = sheet_name.split("]", 1)[1]
                    val = ranges_obj.value
                    if hasattr(val, "tolist"):
                        val = val.tolist()
                    while isinstance(val, list) and len(val) > 0:
                        val = val[0]
                    if val is not None and not isinstance(val, str):
                        cached[(sheet_name.upper(), cell_part.upper())] = val

                logger.info(
                    f"Formula eval: {len(cached)} formula cell(s) computed "
                    f"across {len(self.wb.sheetnames)} sheet(s)"
                )

                if not cached:
                    # No formulas to cache — return as-is
                    with open(tmp.name, "rb") as f:
                        return f.read()

                # Step 3: Patch xlsx XML — inject <v> elements into formula cells
                ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                ns_map = {"ss": ns}

                # Read workbook.xml to map sheet IDs to names
                tmp_out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                tmp_out.close()
                try:
                    with zipfile.ZipFile(tmp.name, "r") as zin:
                        # Map rId → sheet name from workbook.xml
                        wb_xml = etree.fromstring(zin.read("xl/workbook.xml"))
                        sheet_names_by_idx: dict[int, str] = {}
                        for idx, sheet_el in enumerate(wb_xml.findall(f".//{{{ns}}}sheet"), start=1):
                            sheet_names_by_idx[idx] = sheet_el.get("name", "")

                        with zipfile.ZipFile(tmp_out.name, "w", zipfile.ZIP_DEFLATED) as zout:
                            for item in zin.infolist():
                                data = zin.read(item.filename)

                                if (
                                    item.filename.startswith("xl/worksheets/sheet")
                                    and item.filename.endswith(".xml")
                                ):
                                    # Extract sheet index from filename (sheet1.xml → 1)
                                    sheet_idx_str = item.filename.replace(
                                        "xl/worksheets/sheet", ""
                                    ).replace(".xml", "")
                                    try:
                                        sheet_idx = int(sheet_idx_str)
                                    except ValueError:
                                        zout.writestr(item, data)
                                        continue
                                    sheet_name = sheet_names_by_idx.get(sheet_idx, "")
                                    sheet_upper = sheet_name.upper()

                                    tree = etree.fromstring(data)
                                    patched = False
                                    for c_el in tree.iter(f"{{{ns}}}c"):
                                        f_el = c_el.find(f"{{{ns}}}f")
                                        if f_el is None:
                                            continue
                                        ref = (c_el.get("r") or "").upper()
                                        val = cached.get((sheet_upper, ref))
                                        if val is None:
                                            continue
                                        # Inject or update <v> element
                                        v_el = c_el.find(f"{{{ns}}}v")
                                        if v_el is None:
                                            v_el = etree.SubElement(c_el, f"{{{ns}}}v")
                                        v_el.text = str(val)
                                        # Ensure cell type is number (not string)
                                        if c_el.get("t") == "s":
                                            del c_el.attrib["t"]
                                        patched = True

                                    if patched:
                                        data = etree.tostring(
                                            tree,
                                            xml_declaration=True,
                                            encoding="UTF-8",
                                            standalone=True,
                                        )

                                zout.writestr(item, data)

                    with open(tmp_out.name, "rb") as f:
                        return f.read()
                finally:
                    try:
                        os.unlink(tmp_out.name)
                    except OSError:
                        pass

            except Exception as e:
                logger.warning(f"Formula caching failed, returning raw xlsx: {e}")
                with open(tmp.name, "rb") as f:
                    return f.read()

        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _get_sheet(self, sheet_name: str):
        if sheet_name not in self.wb.sheetnames:
            # Auto-create if referenced but not yet created
            ws = self.wb.create_sheet(title=sheet_name)
            self._charts[sheet_name] = []
            return ws
        return self.wb[sheet_name]


def _shift_formula_rows(formula: str, delta: int) -> str:
    """Naively shift absolute row numbers in a formula by delta rows."""
    import re
    if delta == 0:
        return formula

    def replace_ref(m: re.Match) -> str:
        col = m.group(1)
        row = int(m.group(2)) + delta
        return f"{col}{row}"

    # Match cell references like A1, B3, AA10 (no $ prefix for simplicity)
    return re.sub(r"([A-Z]+)(\d+)", replace_ref, formula)
