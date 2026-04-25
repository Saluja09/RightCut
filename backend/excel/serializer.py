"""
RightCut — Workbook → JSON serializer.
Converts an openpyxl Workbook to a WorkbookState dict for the frontend.
"""

from __future__ import annotations

import logging
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from models import CellValue, ChartMeta, SheetState, WorkbookState

logger = logging.getLogger(__name__)


def serialize_workbook(
    wb: openpyxl.Workbook,
    chart_meta: dict[str, list[ChartMeta]] | None = None,
) -> dict:
    """Convert the entire workbook to a frontend-consumable dict."""
    chart_meta = chart_meta or {}
    sheets = []

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            sheet = _serialize_sheet(ws, chart_meta.get(sheet_name, []))
            sheets.append(sheet.model_dump())
        except Exception as e:
            logger.error(f"Failed to serialize sheet '{sheet_name}': {e}")

    active = wb.active.title if wb.active else (sheets[0]["name"] if sheets else "")

    return WorkbookState(sheets=sheets, active_sheet=active).model_dump()


def _serialize_sheet(ws, charts: list[ChartMeta]) -> SheetState:
    """Serialize a single worksheet."""
    max_row = ws.max_row or 1
    raw_max_col = ws.max_column or 1

    # Determine the real data extent: the last column with a non-empty header.
    # Columns beyond this are typically chart-helper columns the agent wrote
    # (e.g. duplicate Category/Percentage columns for chart References) and
    # should NOT appear in the frontend grid.
    max_col = 0
    for col_idx in range(1, raw_max_col + 1):
        hdr = ws.cell(row=1, column=col_idx).value
        if hdr is not None and str(hdr).strip() != "":
            max_col = col_idx
    if max_col == 0:
        max_col = raw_max_col  # fallback: no headers at all, keep everything

    # Collect merged cell ranges so we can handle them
    merged: set[str] = set()
    for merge_range in ws.merged_cells.ranges:
        for cell in merge_range.cells:
            r, c = cell
            ref = f"{get_column_letter(c)}{r}"
            merged.add(ref)

    # Row 1 = headers
    headers: list[str] = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        headers.append(str(cell.value) if cell.value is not None else "")

    # Remaining rows = data
    rows: list[list[CellValue]] = []
    for row_idx in range(2, max_row + 1):
        row_cells: list[CellValue] = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_cells.append(_serialize_cell(cell))
        rows.append(row_cells)

    return SheetState(
        name=ws.title,
        headers=headers,
        rows=rows,
        charts=charts,
        frozen_rows=1,
    )


def _serialize_cell(cell) -> CellValue:
    """Serialize a single openpyxl cell."""
    value = cell.value
    formula: str | None = None
    display_value: Any = None

    if value is None:
        display_value = None
    elif isinstance(value, str) and value.startswith("="):
        formula = value
        display_value = value  # show formula string in grid
    elif isinstance(value, (int, float)):
        display_value = value
    else:
        display_value = str(value)

    # Extract comment text safely
    comment_text: str | None = None
    if cell.comment:
        try:
            comment_text = str(cell.comment.text)
        except Exception:
            comment_text = None

    # Style properties
    bold = False
    bg_color: str | None = None
    font_color: str | None = None

    try:
        if cell.font:
            bold = bool(cell.font.bold)
            if cell.font.color and cell.font.color.type == "rgb":
                font_color = cell.font.color.rgb
    except Exception:
        pass

    try:
        if cell.fill and cell.fill.fill_type not in (None, "none"):
            fg = cell.fill.fgColor
            if fg and fg.type == "rgb" and fg.rgb != "00000000":
                bg_color = fg.rgb
    except Exception:
        pass

    return CellValue(
        value=display_value,
        formula=formula,
        comment=comment_text,
        number_format=cell.number_format if cell.number_format != "General" else None,
        bold=bold,
        bg_color=bg_color,
        font_color=font_color,
    )
